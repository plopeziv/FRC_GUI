import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
 

class ExcelManager: 
    def __init__(self, file_path=None):
        self.file_path = file_path;
        self.job_number = None
        self.job_name = None
        self.job_address = None
        self.dataframe = None;
        self.materials = [];
        self.material_map = {}
        self.labor_map = {}
        self.nte = None
        self.headers = []
        self.data_rows = [];
        self.header_row = 13
        
    def load(self):
        if not self.file_path:
            raise ValueError("No file provided!")
            
        path = Path(self.file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
            
        raw_df = pd.read_excel(path, engine=self.get_excel_engine(), header=None)
        
        self.dataframe = raw_df.iloc[:, 6:]
        
        self.job_number = self.dataframe.iloc[1,8]
        self.job_name = self.dataframe.iloc[2,8]
        self. job_address = self.dataframe.iloc[3,8]
        self.nte = self.dataframe.iloc[1, 5]

        self.get_headers()
        
        self.get_labor_map()
        
        # Can eventually be removed using get_material_map
        self.get_materials()
        
        self.get_material_map()
        
        self.get_data_rows()
        
        return self.dataframe
    
    def get_excel_engine(self):
        extension = Path(self.file_path).suffix.lower()
        
        if extension == ".xlsx":
            engine = "openpyxl"
            
        elif extension == ".xls":
            engine = "xlrd"
            
        else:
            raise ValueError(f"Unsupported Excel file extension: {extension}")
            
        return engine
    
    def get_headers(self):
        """Extract all headers from the header row."""
        if self.dataframe is None:
            raise ValueError("Excel file not loaded yet. Call load() first.")
            
        header_values = self.dataframe.iloc[self.header_row]
        self.headers = [str(material).strip() for material in header_values]
        
        # Get headers because header row doesn't always calculate formulas.
        materials = self.get_materials()
        replacement_index = len(materials)
        self.headers[-replacement_index:] = materials
        
        return self.headers
    
    def get_materials(self, start_col=67, header_row = 6):
        if self.dataframe is None:
            raise ValueError("Excel file not loaded yet. Call load() first.")
            
        row_values = self.dataframe.iloc[header_row, start_col:]
        self.materials = [str(material).strip() for material in row_values if pd.notna(material)]
        
        return self.materials
    
    def get_material_map(self, start_col=67, header_row = 6):

        material_row_values = self.dataframe.iloc[header_row, start_col:]
        price_row_values = self.dataframe.iloc[header_row + 4, start_col:]

        for col_idx, material in enumerate(material_row_values):
            if pd.notna(material):
                material_name = str(material).strip()
                material_price = self._safe_float(price_row_values.iloc[col_idx])
                material_units = self.dataframe.iloc[header_row - 2, (col_idx+ start_col)]
                
                self.material_map[material_name] = {"Sell Per Unit": material_price, "Units": material_units}
        
        return
    
    def get_row_materials(self, check_row, header_row=6, start_col=67):
        job_row = self.dataframe.iloc[check_row, start_col:]

        materials_to_add = []
        for col_idx, quantity in enumerate(job_row):
            if pd.notna(quantity):
                idx_df = start_col + col_idx
                
                material = str(self.dataframe.iloc[header_row, idx_df]).strip()
                units = self.dataframe.iloc[header_row-2, idx_df]
                sell_price = self.dataframe.iloc[header_row+4, idx_df]

                materials_to_add.append({"material": material, "quantity": quantity, "units": units, "sell price": sell_price})

                print(f"material:{material}, quantity: {quantity}, units: {units}, sell price: {sell_price}")
        
        return materials_to_add


    
    def get_labor_map(self):
        rt = self._safe_float(self.dataframe.iloc[10, 12])
        ot = self._safe_float(self.dataframe.iloc[10, 13])
        dt = self._safe_float(self.dataframe.iloc[10, 14])
        
        rt, ot, dt = round(rt, 2), round(ot, 2), round(dt, 2)
    
        self.labor_map = {
            "RT": {"rate": rt},
            "OT": {"rate": ot},
            "DT": {"rate": dt},
            "OT DIFF": {"rate": round(ot-rt, 2)},
            "DT DIFF": {"rate": round(dt-rt, 2)},
            }

    
    def get_data_rows (self):
        ticket_number_col = 8
        
        for index in range(self.header_row+1, len(self.dataframe)):
            data_row = self.dataframe.iloc[index]
            
            if self.is_row_empty(data_row):
                break

            #Coerce index/ticket_number into a string 
            if pd.notna(data_row[ticket_number_col]):
                data_row[ticket_number_col] = str(data_row[ticket_number_col]).strip()
            
            self.data_rows.append(data_row)
            
    def is_row_empty(self, row):
        return not any(bool(cell) for cell in row.fillna(0))
    
    def insert_ticket (self, frc_ticket):
        path = self.ensure_xlsx_copy(self.file_path)
           
        ws, wb = self._open_worksheet(path, "TICKET TRACKING")
        

        
        new_row = self.find_ticket_row_ws(ws, frc_ticket["Ticket Number"])
        
        self._insert_ticket_info(ws, new_row, frc_ticket)
        
        
        self._insert_labor(ws, new_row, frc_ticket["Labor"])
        
        self._insert_materials(ws, new_row, frc_ticket["Materials"])
        
        wb.save(path)
        wb.close()
        
    def _safe_float(self, incoming_value):
        try:
            incoming_value = str(incoming_value).strip().lower()
            return float(incoming_value) if incoming_value not in ("none", "", "nan") else 0.0
        except(ValueError, TypeError):
            return 0.0
        
    def _open_worksheet(self, path, tab_title):
        wb = load_workbook(path)
        
        if tab_title not in wb.sheetnames:
            wb.close()
            raise ValueError(f"The Excel file does not contain a '{tab_title}' sheet.")
            
        return wb[tab_title], wb
        
    def _insert_ticket_info(self, worksheet, ticket_row, ticket_object):
        worksheet.cell(row=ticket_row, column=7, value=ticket_object["Date"])
        worksheet.cell(row=ticket_row, column=8, value=ticket_object["Signature"])
        worksheet.cell(row=ticket_row, column=9, value=ticket_object["Ticket Number"])
        worksheet.cell(row=ticket_row, column=10, value=ticket_object["Type"])
        worksheet.cell(row=ticket_row, column=11, value=ticket_object["Description"])
        
    def _insert_labor(self, worksheet, ticket_row, labor_object):
        worksheet.cell(row=ticket_row, column=19, value = self._safe_float(labor_object["RT"]["hours"]))
        worksheet.cell(row=ticket_row, column=20, value = self._safe_float(labor_object["OT"]["hours"]))
        worksheet.cell(row=ticket_row, column=21, value = self._safe_float(labor_object["DT"]["hours"]))
        worksheet.cell(row=ticket_row, column=22, value = self._safe_float(labor_object["OT DIFF"]["hours"]))
        worksheet.cell(row=ticket_row, column=23, value = self._safe_float(labor_object["DT DIFF"]["hours"]))
        
    def _insert_materials(self, worksheet, material_row, materials):
        for material_object in materials:
            material_name = material_object["material"].strip()
            material_quantity = self._safe_float(material_object["quantity"])
                
            try: 
                column_index = self.headers.index(material_name) + 7
                
                worksheet.cell(row=material_row, column=column_index, value=material_quantity)
                
            except ValueError:
                print(F"{material_name} not found in headers.... Skipped!")

        
        
    def ensure_xlsx_copy(self, path):
        """
        If the working copy path is .xls, convert it to .xlsx for openpyxl.
        Returns the new path (string) compatible with openpyxl.
        """
        path = Path(path)
        if path.suffix.lower() == ".xls":
            new_path = path.with_suffix(".xlsx")
            if not new_path.exists():
                print(f"⚙️ Converting {path.name} → {new_path.name}")
                df = pd.read_excel(path, engine="xlrd", header=None)
                df.to_excel(new_path, index=False, header=False)
            return str(new_path)
        return str(path)
    
    def find_ticket_row_ws(self, worksheet, ticket_number):
        ticket_number = str(ticket_number).strip()        
        
        row = self.header_row + 2
        
        while True:
            cell_value = worksheet.cell(row=row, column=9).value
            
            if cell_value in (None, "", " "):
                return row
            
            if str(cell_value).strip() == ticket_number:
                return row

            row += 1

    def find_ticket_row_df(self, ticket_number):
        # DF and WS rows are the same. Use this function for outside of the worksheet.

        if self.dataframe is None:
            raise ValueError("Excel file not loaded yet. Call load() first.")

        ticket_number = str(ticket_number).strip()

        row = self.header_row + 2

        while row < len(self.dataframe):
            print(row)
            cell_value = self.dataframe.loc[row, 8]

            if str(cell_value).strip() == ticket_number:
                return row
            
            row += 1
        
        return row
        
    
if __name__ =="__main__":
    test_path = "1234 - DETAILED TICKET LISTING (1-7-26).xlsx"
    
    manager = ExcelManager(test_path)
    
    df = manager.load()
    
    incoming_ticket = {
      'Job Number': '123456',
      'Job Name': 'Test Job',
      'Ticket Number': '00001',
      'Job Address': '1234 Fake Street',
      'Date': '11/10/25',
      'Signature': 'Yes',
      'Type': 'REGULAR',
      'Description': 'Fixed pump housing leak',
      'Labor': {
          'RT': {'hours':'12', 'rate': '153.15'}, 
          'OT': {'hours':'4', 'rate': '197.70'}, 
          'DT': {'hours':'0', 'rate': '237.72'}, 
          'OT DIFF': {'hours':'0', 'rate':'40.55'}, 
          'DT DIFF': {'hours': '0', 'rate': '80.57'}
          },
      'Materials': [
          {
              'material': '1/4 UNDERLAYMENT 4 X 5"', 
              'quantity': '3',
              'units': 'EA',
              'sell price': '53.44',
          }, 
          {
              'material': 'MAPEI QUICK PATCH 25LB', 
              'quantity': '10',
              'units': 'BG',
              'sell price': '34.87'
          },
          {
              'material': 'HEPA SANDER#302 & VAC #701', 
              'quantity': '1',
              'units': 'EA',
              'sell price': '150'
          }
     ]
    }
    
    # manager.insert_ticket(incoming_ticket)
        