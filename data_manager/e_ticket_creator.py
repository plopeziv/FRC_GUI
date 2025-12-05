# -*- coding: utf-8 -*-
"""
Created on Sat Nov 22 12:04:44 2025

@author: plopez
"""
from pathlib import Path

class ETicketCreator:
    def __init__(self, file_path=None, incoming_ticket=None):
        self.file_path = file_path;
        self.workbook = None
        self.incoming_ticket = incoming_ticket;
        
    def load_ticket(self):
        from openpyxl import load_workbook
        
        if not self.file_path:
            raise ValueError("No folder provided!")
            
        path = Path(self.file_path) / "E-ticket Replacement EDITABLE - PYTHON.xlsx"
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        
        path = self._ensure_xlsx_copy(path)
        wb = load_workbook(path)
        ws = wb.active
        
        self._insert_job_info(ws)
        
        self._unmerge_from_row(ws, start_row=32)

        self._insert_labor(ws)
        
        self._insert_materials(ws)
        
        self._calculate_ticket_total(ws)
        
        self._format_footer(ws)
        
        save_path =  Path(self.file_path) / f"{self.incoming_ticket['Job Number']} - {self.incoming_ticket['Ticket Number']}.xlsx"
        
        wb.save(save_path)
        
    
    def _ensure_xlsx_copy(self, path):
        """
        If the working copy path is .xls, convert it to .xlsx for openpyxl.
        Returns the new path (string) compatible with openpyxl.
        """
        import pandas as pd

        if path.suffix.lower() == ".xls":
            new_path = path.with_suffix(".xlsx")
            if not new_path.exists():
                print(f"⚙️ Converting {path.name} → {new_path.name}")
                df = pd.read_excel(path, engine="xlrd", header=None)
                df.to_excel(new_path, index=False, header=False)
            return str(new_path)
        return str(path)
    
    def _get_excel_engine(self):
        extension = Path(self.file_path).suffix.lower()
        
        if extension == ".xlsx":
            engine = "openpyxl"
            
        elif extension == ".xls":
            engine = "xlrd"
            
        else:
            raise ValueError(f"Unsupported Excel file extension: {extension}")
            
        return engine
    
    def _insert_job_info(self, ws):
        from datetime import datetime
        import pandas as pd
        
        ws['B4'] = self.incoming_ticket["Job Number"]
        ws['B5'] = self.incoming_ticket["Job Name"]
        ws['B6'] = self.incoming_ticket["Job Address"]
        ws['B7'] = self.incoming_ticket["Installers"]
        ws['B8'] = self.incoming_ticket["Work Location"]
        
        ws['H4']= self.incoming_ticket["Ticket Number"]
        ws['H5']= pd.to_datetime(datetime.today(), format="%m/%d/%y", errors="coerce")
            
        ws['H6']= pd.to_datetime(self.incoming_ticket["Date"], format="%m/%d/%y", errors="coerce")        
        
        ws['B10']= self.incoming_ticket["Description"]
        
    def _insert_labor(self, ws):
        labor_object = self.incoming_ticket["Labor"]
        
        category_map = {
        "RT": "REGULAR TIME",
        "OT": "OVERTIME",
        "DT": "DOUBLE TIME",
        "OT DIFF": "OVERTIME DIFF",
        "DT DIFF": "DOUBLE TIME DIFF",
        }
        
        # Keep only catagories with hours > 0 and translate their keys
        active_labor = {
            category_map[k]: v for k, v 
            in labor_object.items()
            if self._safe_float(v["hours"]) > 0}
        
        #Safety Check
        if not active_labor:
            rt_object = labor_object["RT"]
            active_labor = {
                    "REGULAR TIME": {
                        "hours": "0",
                        "rate": rt_object["rate"]
                        }
                }
        
        iterator = iter(active_labor)
        first_key = next(iterator)
        
        # Write first labor entry into row 17
        ws['B17'] = round(self._safe_float(active_labor[first_key]["hours"]), 1)
        ws['C17'] = first_key
        ws['F17'] = round(self._safe_float(active_labor[first_key]["rate"]),2)
        ws['I17'] = "=B17*F17"
        
        # Insert Additional Labor Rows
        current_row = 17
        for key in iterator:
           current_row += 1
           ws.insert_rows(current_row)
           
           self._copy_and_insert_row(ws, 17, current_row)
           
           ws.merge_cells(start_row=current_row, end_row=current_row, 
                          start_column=3, end_column=4)
           
           ws[f'B{current_row}'] = round(self._safe_float(active_labor[key]["hours"]), 1)
           ws[f'C{current_row}'] = key
           ws[f'F{current_row}'] = round(self._safe_float(active_labor[key]["rate"]),2)
           ws[f'I{current_row}'] = f'=B{current_row}*F{current_row}'
           
        ws[f'I{current_row+2}'] = f'=SUM(I17:I{current_row})'
           
        
        
    def _copy_and_insert_row(self, worksheet, src_row, dest_row, max_col=9):
        from copy import copy
        
        for col in range(1, max_col+1):
            src_cell = worksheet.cell(row=src_row, column=col)
            dest_cell = worksheet.cell(row=dest_row, column=col)
            
            dest_cell.value = src_cell.value
            
            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.number_format = src_cell.number_format
                dest_cell.protection = copy(src_cell.protection)
                dest_cell.alignment = copy(src_cell.alignment)
                
    def _insert_materials(self, ws):
        material_object = self.incoming_ticket["Materials"]
        
        if not material_object:
           total_material_row = self._find_material_row(ws, "Total Material", column="G")
           ws[f'I{total_material_row}'] = 0
           return
       
        start_row = self._find_material_row(ws) + 1
        
        if start_row is None:
            raise ValueError("Material starting row not found")
            
        # First Material Row
        ws[f'B{start_row}'] = material_object[0]["quantity"]
        ws[f'C{start_row}'] = material_object[0]["material"]
        ws[f'F{start_row}'] = round(self._safe_float(material_object[0]["sell price"]))
        ws[f'I{start_row}'] = f'=B{start_row} * F{start_row}'
        
        ws.row_dimensions[start_row].height = 25
        
        ws.merge_cells(start_row=start_row, end_row=start_row, 
                       start_column=3, end_column=4)
      
        # Any Material After 1
        current_row = start_row
        
        for material in material_object[1:]:
            current_row += 1
            ws.insert_rows(current_row)
            
            self._copy_and_insert_row(ws, start_row, current_row)
            
            ws[f'B{current_row}'] = material["quantity"]
            ws[f'C{current_row}'] = material["material"]
            ws[f'F{current_row}'] = round(self._safe_float(material["sell price"]))
            ws[f'I{current_row}'] = f'=B{current_row} * F{current_row}'
            
            ws.row_dimensions[current_row].height = 25
            
            ws.merge_cells(start_row=current_row, end_row=current_row, 
                           start_column=3, end_column=4)
        
        
        # Create the material total summary
        total_material_row = self._find_material_row(ws, "Total Material", column="G")
        ws[f'I{total_material_row}'] = f'=SUM(I{start_row}:I{current_row})'
        
        
        
    def _calculate_ticket_total(self, ws):
        labor_total_row = self._find_material_row(ws, "Total Hours", column='G')
        material_total_row = self._find_material_row(ws, "Total Material", column='G')
        ticket_total_row = self._find_material_row(ws, "Total Ticket", column='G')
        
        formula = f'=I{labor_total_row}+I{material_total_row}'
        ws[f'I{ticket_total_row}'] = formula
        ws.row_dimensions[ticket_total_row].height = 15
        

        
    def _find_material_row(self, ws, search_term="Material Used:", column="A"):
        for row in range(1,50):
            check_value = ws[f'{column}{row}'].value
            if check_value and search_term in str(check_value):
                return row
        
        return None

    def _format_footer(self, ws):
        
        # Find the row you want to merge cells for
        ws_row = self._find_material_row(ws, search_term="Work Status")
        
        self._unmerge_row(ws, ws_row)
                
        # Merge columns 1–2 (A–B)
        ws.merge_cells(start_row=ws_row, end_row=ws_row, start_column=1, end_column=2)
        
        # Merge columns 3–6 (C–F)
        ws.merge_cells(start_row=ws_row, end_row=ws_row, start_column=3, end_column=6)
        
        self._add_bottom_border(ws, ws_row, 7, 9)
        
        ws.row_dimensions[ws_row].height = 15
        
        # as Row
        as_row = self._find_material_row(ws, search_term="Authorization Status")
        
        self._unmerge_row(ws, as_row)
                
        ws.merge_cells(start_row=as_row, end_row=as_row, start_column=1, end_column=2)
        
        ws.merge_cells(start_row=as_row, end_row=as_row, start_column=3, end_column=9)
        ws.merge_cells(start_row=as_row+1, end_row=as_row+1, start_column=3, end_column=9)
        
        ws.row_dimensions[as_row].height = 15
        ws.row_dimensions[as_row+1].height = 15
        ws.row_dimensions[as_row+2].height = 15
        
        # fs_row
        fs_row = self._find_material_row(ws, search_term="Field Supervisor")
        
        for row in range(ws_row - 1, fs_row):
            ws.row_dimensions[row].height = 15
        
        self._unmerge_row(ws, fs_row)
        
        ws.merge_cells(start_row=fs_row, end_row=fs_row, 
                       start_column=1, end_column=2)
        
        self._add_bottom_border(ws, fs_row, 3, 4)
        self._add_bottom_border(ws, fs_row, 7, 9)
        
        ws.row_dimensions[fs_row].height = 25
       
        # pm_row
        pm_row = self._find_material_row(ws, search_term="Project Manager")
        self._unmerge_row(ws, pm_row)
        ws.merge_cells(start_row=pm_row, end_row=pm_row, 
                       start_column=1, end_column=2)
        
        self._add_bottom_border(ws, pm_row, 3, 4)
        self._add_bottom_border(ws, pm_row, 7, 9)
        
        ws.row_dimensions[pm_row].height = 25
        
    def _unmerge_row(self, ws, row_number):

        # Make a copy of merged ranges because we can't modify the list while iterating
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row <= row_number <= merged_range.max_row:
                ws.unmerge_cells(str(merged_range))
                
    def _unmerge_from_row(self, ws, start_row):
        """
        Unmerge all merged cells from a given row downward.
        """
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row >= start_row:
                ws.unmerge_cells(str(merged_range))
                
    def _add_bottom_border(self, ws, row, start_col, end_col, style='thin'):
        
        from openpyxl.styles import Border, Side
        
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=Side(style=style)
            )
            
    def _safe_float(self, incoming_value):
        try:
            incoming_value = str(incoming_value).strip().lower()
            return float(incoming_value) if incoming_value not in ("none", "", "nan") else 0.0
        except(ValueError, TypeError):
            return 0.0

        
        
    
        
        
        
    
    
        
    

if __name__ =="__main__":
    test_path = "E-ticket Replacement EDITABLE - PYTHON.xlsx"
    
    
    incoming_ticket = {
      'Job Number': '123456',
      'Job Name': 'Test Job',
      'Ticket Number': '00003',
      'Job Address': '1234 Fake Street',
      'Date': '11/10/25',
      'Signature': 'Yes',
      'Type': 'REGULAR',
      'Installers': 'Juan',
      'Work Location': '35TH FLR',
      'Description': 'Fixed pump housing leak',
      'Labor': {
          'RT': {'hours':'12', 'rate': '153.15'}, 
          'OT': {'hours':'4', 'rate': '197.70'}, 
          'DT': {'hours':'0', 'rate': '237.72'}, 
          'OT DIFF': {'hours':'0', 'rate':'40.55'}, 
          'DT DIFF': {'hours': '0', 'rate': '80.57'}
          },
      'Materials': [
          {
              'material': 'MAPEI PLANIPREP SC 10LB BAG', 
              'quantity': '3',
              'sell price': '34.35'
          }, 
          {
              'material': 'MAPEI QUICK PATCH 25LB', 
              'quantity': '10',
              'sell price': '34.87'
          },
          {
              'material': 'HEPA SANDER#302 & VAC #701', 
              'quantity': '1',
              'sell price': '150'
          }
     ]
    }
    
    wb = ETicketCreator(test_path, incoming_ticket).load_ticket()
    